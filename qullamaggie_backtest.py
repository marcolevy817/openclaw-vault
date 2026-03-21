"""
Qullamaggie Breakout Strategy — Walk-Forward Backtester
Uses the same scanner logic from qullamaggie_scanner.py
Outputs: performance stats + detailed trade_log.xlsx

Run:
    python3 qullamaggie_backtest.py --years 3
    python3 qullamaggie_backtest.py --symbol NVDA --years 5
    python3 qullamaggie_backtest.py --years 3 --capital 100000 --output trades.xlsx
"""

import argparse
import io
import math
import os
import sys
import warnings
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from tabulate import tabulate

warnings.filterwarnings("ignore")

# ── Import scanner helpers ─────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from qullamaggie_scanner import (
    compute_adr,
    check_ma_structure,
    detect_consolidation,
    compute_setup_score,
    CONFIG,
    KNOWN_BIOTECH,
)

ET = ZoneInfo("America/New_York")

# ── Backtest config ────────────────────────────────────────────────────────
BT_CONFIG = {
    "starting_capital": 100_000,
    "risk_pct": 0.005,          # 0.5% per trade
    "max_position_pct": 0.30,   # max 30% per position
    "max_open_positions": 10,   # concurrent positions cap
    "slippage_pct": 0.001,      # 0.1% slippage on fills
    "commission_per_share": 0.005,  # $0.005/share (Interactive Brokers-style)
    "partial_exit_day": 3,
    "partial_exit_frac": 0.40,
    "early_partial_r_multiple": 2.0,
    "trail_ma": 10,
    "spy_sma50_tolerance": 0.97,
    "nasdaq_cache_path": "/tmp/nasdaq_universe.csv",
    "nasdaq_url": "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=10000&exchange=nasdaq",
    "nyse_url": "https://api.nasdaq.com/api/screener/stocks?tableonly=true&limit=10000&exchange=nyse",
}


# ──────────────────────────────────────────────────────────────────────────
# NASDAQ / NYSE UNIVERSE — clean common stocks only
# ──────────────────────────────────────────────────────────────────────────
def fetch_nasdaq_universe(force_refresh: bool = False) -> list:
    """
    Download full Nasdaq + NYSE + AMEX stock list (~6,600 clean common stocks).
    Uses local cache file if fresh (< 24h). Falls back to workspace file.
    Returns list of ticker symbols — alpha-only, 1–5 chars (no warrants/units/preferred).
    """
    cache = BT_CONFIG["nasdaq_cache_path"]
    workspace_cache = "/root/.openclaw/workspace/us_universe.txt"

    # Check disk cache
    if not force_refresh and os.path.exists(cache):
        age_hours = (datetime.now().timestamp() - os.path.getmtime(cache)) / 3600
        if age_hours < 24:
            with open(cache) as f:
                symbols = [l.strip() for l in f if l.strip()]
            print(f"[UNIVERSE] Loaded {len(symbols)} symbols from cache (age: {age_hours:.1f}h)")
            return symbols

    # Check workspace pre-built file
    if not force_refresh and os.path.exists(workspace_cache):
        with open(workspace_cache) as f:
            symbols = [l.strip() for l in f if l.strip()]
        print(f"[UNIVERSE] Loaded {len(symbols)} symbols from workspace cache")
        return symbols

    print("[UNIVERSE] Fetching fresh universe from Nasdaq screener API...")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.nasdaq.com/",
    }

    all_syms = []
    for exchange in ["nasdaq", "nyse", "amex"]:
        url = (f"https://api.nasdaq.com/api/screener/stocks"
               f"?tableonly=true&limit=5000&offset=0&exchange={exchange}&download=true")
        try:
            resp = requests.get(url, headers=headers, timeout=20)
            data = resp.json()
            rows = data.get("data", {}).get("rows", [])
            print(f"  {exchange.upper()}: {len(rows)} rows")
            for row in rows:
                sym = row.get("symbol", "").strip()
                if sym and sym.isalpha() and 1 <= len(sym) <= 5:
                    all_syms.append(sym)
        except Exception as e:
            print(f"  [WARN] {exchange}: {e}")

    symbols = sorted(set(all_syms))
    print(f"[UNIVERSE] {len(symbols)} clean US common stock symbols")

    # Save both caches
    with open(cache, "w") as f:
        f.write("\n".join(symbols))
    with open(workspace_cache, "w") as f:
        f.write("\n".join(symbols))

    return symbols


# ──────────────────────────────────────────────────────────────────────────
# BULK DATA DOWNLOAD
# ──────────────────────────────────────────────────────────────────────────
def download_all_bars(symbols: list, years: int = 3) -> dict:
    """
    Batch download OHLCV bars for all symbols.
    Returns dict: {symbol: DataFrame}.
    Uses chunked yfinance downloads for speed.
    """
    end = datetime.now(ET)
    start = end - timedelta(days=int(years * 365) + 60)
    start_str = start.strftime("%Y-%m-%d")
    end_str = end.strftime("%Y-%m-%d")

    print(f"[DATA] Downloading {len(symbols)} symbols ({years}y: {start_str} → {end_str})")
    print("       This may take several minutes...")

    chunk_size = 100
    all_data = {}
    chunks = [symbols[i:i + chunk_size] for i in range(0, len(symbols), chunk_size)]

    for idx, chunk in enumerate(chunks):
        pct = (idx / len(chunks)) * 100
        print(f"  Chunk {idx+1}/{len(chunks)} ({pct:.0f}%) — {len(all_data)} symbols loaded so far")
        try:
            raw = yf.download(
                tickers=" ".join(chunk),
                start=start_str,
                end=end_str,
                interval="1d",
                group_by="ticker",
                auto_adjust=True,
                progress=False,
                threads=True,
            )
        except Exception as e:
            print(f"  [WARN] Chunk {idx+1} failed: {e}")
            continue

        for sym in chunk:
            try:
                if sym in raw.columns.get_level_values(0):
                    df = raw[sym][["Open", "High", "Low", "Close", "Volume"]].copy()
                    df.dropna(inplace=True)
                    if len(df) >= 65:  # minimum bars needed
                        all_data[sym] = df
            except Exception:
                pass

    print(f"[DATA] {len(all_data)} symbols with sufficient history")
    return all_data


# ──────────────────────────────────────────────────────────────────────────
# SINGLE-SYMBOL SCANNER (given a slice of historical data)
# ──────────────────────────────────────────────────────────────────────────
def scan_symbol_on_date(df_slice: pd.DataFrame) -> dict | None:
    """
    Run the full Qullamaggie scan on a historical slice (df_slice = data up to 'today').
    Returns a setup dict if all criteria pass, else None.
    """
    if len(df_slice) < 65:
        return None

    # Hard filters on this slice
    last_close = df_slice["Close"].iloc[-1]
    if last_close < CONFIG["min_price"]:
        return None

    avg_dollar_vol = (df_slice["Close"] * df_slice["Volume"]).rolling(20).mean().iloc[-1]
    if avg_dollar_vol < CONFIG["min_avg_dollar_vol"]:
        return None

    adr_series = compute_adr(df_slice)
    if pd.isna(adr_series.iloc[-1]) or adr_series.iloc[-1] < CONFIG["min_adr_pct"]:
        return None

    # Momentum
    if len(df_slice) < 65:
        return None
    move_21d = (df_slice["Close"].iloc[-1] - df_slice["Close"].iloc[-22]) / df_slice["Close"].iloc[-22] * 100
    move_63d = (df_slice["Close"].iloc[-1] - df_slice["Close"].iloc[-64]) / df_slice["Close"].iloc[-64] * 100
    if not (move_21d >= CONFIG["min_move_21d"] or move_63d >= CONFIG["min_move_63d"]):
        return None

    # MA structure
    ma_ok, dist50 = check_ma_structure(df_slice)
    if not ma_ok:
        return None

    # Consolidation
    consol_ok, N, consol_high, consol_low, qm = detect_consolidation(df_slice)
    if not consol_ok or not qm:
        return None
    if not qm.get("surfing_ok", True):
        return None

    score = compute_setup_score(qm, dist50, 0)  # sector bonus skipped in backtest for speed
    adr_val = adr_series.iloc[-1]

    return {
        "score": score,
        "consol_high": consol_high,
        "consol_low": consol_low,
        "consol_days": N,
        "adr_pct": adr_val,
        "dist_from_50": dist50,
        "consol_range_pct": qm["consolidation_range_pct"],
        "volume_ratio": qm["volume_ratio"],
        "higher_lows": qm["higher_lows"],
        "move_21d": round(move_21d, 2),
        "move_63d": round(move_63d, 2),
        "baseline_vol": qm.get("baseline_vol", 0),
    }


# ──────────────────────────────────────────────────────────────────────────
# WALK-FORWARD BACKTEST ENGINE
# ──────────────────────────────────────────────────────────────────────────
def run_backtest(all_bars: dict, starting_capital: float, years: int) -> tuple:
    """
    Walk-forward simulation.
    For each trading day:
      1. Run scanner on all symbols (using only past data)
      2. Enter breakouts (price > consolidation_high on next open)
      3. Manage open positions (partials, trailing stop, hard stop)

    Returns (trade_log: list[dict], equity_curve: pd.Series)
    """
    # Align all data on common trading days using SPY as the calendar
    spy_bars = all_bars.get("SPY")
    if spy_bars is None:
        # Fallback: use the symbol with the most bars
        spy_bars = max(all_bars.values(), key=len)

    trading_days = spy_bars.index.tolist()

    # Trim to the backtest window
    cutoff_start = spy_bars.index[-1] - pd.DateOffset(years=years)
    trading_days = [d for d in trading_days if d >= cutoff_start]

    print(f"\n[BACKTEST] Simulating {len(trading_days)} trading days | "
          f"{len(all_bars)} symbols | Capital: ${starting_capital:,.0f}")

    equity = starting_capital
    cash = starting_capital
    open_positions = {}   # symbol → position dict
    trade_log = []
    equity_curve = {trading_days[0]: equity}

    # Need enough warm-up bars before simulation starts
    warmup = 200

    for day_idx, today in enumerate(trading_days):
        if day_idx % 100 == 0:
            pct = day_idx / len(trading_days) * 100
            print(f"  Day {day_idx}/{len(trading_days)} ({pct:.0f}%) | "
                  f"Equity: ${equity:,.0f} | Positions: {len(open_positions)}")

        # ── Manage open positions at today's open / close ──────────────
        positions_to_close = []
        for sym, pos in list(open_positions.items()):
            sym_df = all_bars.get(sym)
            if sym_df is None or today not in sym_df.index:
                continue

            today_loc = sym_df.index.get_loc(today)
            if today_loc < 1:
                continue

            today_open = sym_df["Open"].iloc[today_loc]
            today_close = sym_df["Close"].iloc[today_loc]
            today_low = sym_df["Low"].iloc[today_loc]
            today_high = sym_df["High"].iloc[today_loc]

            entry = pos["entry_price"]
            stop = pos["stop_price"]
            shares = pos["shares"]
            partial_taken = pos.get("partial_taken", False)
            entry_date = pos["entry_date"]
            days_held = (today - entry_date).days

            exit_price = None
            exit_reason = None

            # Hard stop hit at open (gap down through stop)
            if today_open <= stop:
                exit_price = today_open  # fill at open (gap risk)
                exit_reason = "stop_loss_gap"
            # Intraday stop hit
            elif today_low <= stop:
                exit_price = stop
                exit_reason = "stop_loss"
            else:
                # Trail: close below SMA10
                df_slice = sym_df.iloc[:today_loc + 1]
                if len(df_slice) >= 10:
                    sma10 = df_slice["Close"].rolling(10).mean().iloc[-1]
                    if today_close < sma10 and partial_taken:
                        exit_price = today_close
                        exit_reason = "trail_sma10"

                # Partial exit
                if exit_price is None and not partial_taken:
                    risk = entry - pos["initial_stop"]
                    early_trigger = (today_close - entry) >= BT_CONFIG["early_partial_r_multiple"] * risk
                    day_trigger = days_held >= BT_CONFIG["partial_exit_day"] and today_close > entry

                    if early_trigger or day_trigger:
                        sell_qty = math.floor(shares * BT_CONFIG["partial_exit_frac"])
                        if sell_qty > 0:
                            fill = today_close * (1 - BT_CONFIG["slippage_pct"])
                            commission = sell_qty * BT_CONFIG["commission_per_share"]
                            proceeds = fill * sell_qty - commission
                            cost_basis = entry * sell_qty
                            pnl = proceeds - cost_basis

                            trade_log.append({
                                "symbol": sym,
                                "entry_date": entry_date.date(),
                                "exit_date": today.date(),
                                "days_held": days_held,
                                "entry_price": round(entry, 4),
                                "exit_price": round(fill, 4),
                                "shares": sell_qty,
                                "pnl_dollars": round(pnl, 2),
                                "pnl_pct": round((fill / entry - 1) * 100, 2),
                                "exit_reason": "partial_exit",
                                "setup_score": pos.get("score", 0),
                                "adr_pct": round(pos.get("adr_pct", 0), 2),
                                "consol_days": pos.get("consol_days", 0),
                                "consol_range_pct": round(pos.get("consol_range_pct", 0), 2),
                                "volume_ratio": round(pos.get("volume_ratio", 0), 3),
                                "breakout_level": round(pos.get("breakout_level", 0), 4),
                                "prior_move_pct": round(pos.get("move_63d", 0), 2),
                                "initial_risk_pct": round(pos.get("initial_risk_pct", 0), 2),
                            })

                            cash += proceeds
                            equity += pnl
                            pos["shares"] -= sell_qty
                            pos["stop_price"] = entry  # move to breakeven
                            pos["partial_taken"] = True

            if exit_price is not None:
                # Apply slippage
                if exit_reason in ("stop_loss", "stop_loss_gap"):
                    fill = exit_price * (1 - BT_CONFIG["slippage_pct"])
                else:
                    fill = exit_price * (1 - BT_CONFIG["slippage_pct"])

                commission = shares * BT_CONFIG["commission_per_share"]
                proceeds = fill * shares - commission
                cost_basis = entry * shares
                pnl = proceeds - cost_basis

                trade_log.append({
                    "symbol": sym,
                    "entry_date": entry_date.date(),
                    "exit_date": today.date(),
                    "days_held": days_held,
                    "entry_price": round(entry, 4),
                    "exit_price": round(fill, 4),
                    "shares": shares,
                    "pnl_dollars": round(pnl, 2),
                    "pnl_pct": round((fill / entry - 1) * 100, 2),
                    "exit_reason": exit_reason,
                    "setup_score": pos.get("score", 0),
                    "adr_pct": round(pos.get("adr_pct", 0), 2),
                    "consol_days": pos.get("consol_days", 0),
                    "consol_range_pct": round(pos.get("consol_range_pct", 0), 2),
                    "volume_ratio": round(pos.get("volume_ratio", 0), 3),
                    "breakout_level": round(pos.get("breakout_level", 0), 4),
                    "prior_move_pct": round(pos.get("move_63d", 0), 2),
                    "initial_risk_pct": round(pos.get("initial_risk_pct", 0), 2),
                })

                cash += proceeds
                equity += pnl
                positions_to_close.append(sym)

        for sym in positions_to_close:
            del open_positions[sym]

        # ── Market condition check ──────────────────────────────────────
        spy_loc = spy_bars.index.get_loc(today) if today in spy_bars.index else -1
        market_ok = True
        if spy_loc >= 50:
            spy_sma50 = spy_bars["Close"].iloc[spy_loc - 50:spy_loc].mean()
            spy_close = spy_bars["Close"].iloc[spy_loc]
            market_ok = spy_close >= spy_sma50 * BT_CONFIG["spy_sma50_tolerance"]

        # ── Scan for new setups (using only data up to today) ──────────
        if len(open_positions) < BT_CONFIG["max_open_positions"] and cash > 1000:
            today_idx_global = trading_days.index(today)
            if today_idx_global < warmup:
                equity_curve[today] = equity
                continue

            new_setups = []
            for sym, sym_df in all_bars.items():
                if sym == "SPY" or sym in open_positions:
                    continue
                if today not in sym_df.index:
                    continue

                today_loc = sym_df.index.get_loc(today)
                if today_loc < warmup:
                    continue

                # Slice: all data UP TO AND INCLUDING today (no lookahead)
                df_slice = sym_df.iloc[:today_loc + 1]

                setup = scan_symbol_on_date(df_slice)
                if setup is not None:
                    new_setups.append((sym, setup, df_slice))

            # Sort by score, enter highest-quality setups first
            new_setups.sort(key=lambda x: x[1]["score"], reverse=True)

            for sym, setup, df_slice in new_setups:
                if len(open_positions) >= BT_CONFIG["max_open_positions"]:
                    break
                if cash <= 0:
                    break

                # Entry: check if today's price broke above consolidation_high
                # We use today's close > consol_high as the signal
                # (simulates EOD breakout — buy at next open)
                today_close_sym = df_slice["Close"].iloc[-1]
                consol_high = setup["consol_high"]

                if today_close_sym <= consol_high:
                    continue  # not breaking out yet

                # ADR gate: don't chase if moved > 1 ADR from today's low
                today_low_sym = df_slice["Low"].iloc[-1]
                adr_abs = df_slice["Close"].iloc[-2] * (setup["adr_pct"] / 100)
                if today_close_sym > today_low_sym + adr_abs:
                    continue  # too extended

                # Entry at NEXT day's open (simulate buy-at-open)
                today_loc = df_slice.index.get_loc(today)
                remaining_days = [d for d in trading_days if d > today]
                if not remaining_days:
                    continue
                next_day = remaining_days[0]
                next_sym_df = all_bars.get(sym)
                if next_sym_df is None or next_day not in next_sym_df.index:
                    continue

                entry_price = next_sym_df["Open"].loc[next_day] * (1 + BT_CONFIG["slippage_pct"])
                stop_price = next_sym_df["Low"].loc[next_day]  # day-of-entry LOD

                # Stop validation: stop can't exceed 1 ADR wide
                stop_width_pct = (entry_price - stop_price) / entry_price * 100
                if stop_width_pct > setup["adr_pct"]:
                    continue

                if stop_price >= entry_price:
                    continue

                # Position sizing
                risk_per_share = entry_price - stop_price
                dollar_risk = equity * BT_CONFIG["risk_pct"]
                if not market_ok:
                    dollar_risk *= BT_CONFIG.get("weak_market_size_factor", 0.5)

                raw_shares = math.floor(dollar_risk / risk_per_share)
                max_shares = math.floor((equity * BT_CONFIG["max_position_pct"]) / entry_price)
                shares = min(raw_shares, max_shares)

                if shares <= 0:
                    continue

                cost = shares * entry_price + shares * BT_CONFIG["commission_per_share"]
                if cost > cash:
                    shares = math.floor((cash * 0.95) / entry_price)
                    if shares <= 0:
                        continue
                    cost = shares * entry_price + shares * BT_CONFIG["commission_per_share"]

                cash -= cost

                open_positions[sym] = {
                    "entry_price": entry_price,
                    "initial_stop": stop_price,
                    "stop_price": stop_price,
                    "shares": shares,
                    "entry_date": next_day,
                    "partial_taken": False,
                    "score": setup["score"],
                    "adr_pct": setup["adr_pct"],
                    "consol_days": setup["consol_days"],
                    "consol_range_pct": setup["consol_range_pct"],
                    "volume_ratio": setup["volume_ratio"],
                    "breakout_level": consol_high,
                    "move_63d": setup["move_63d"],
                    "initial_risk_pct": round(stop_width_pct, 2),
                }

        # Mark-to-market equity
        pos_value = sum(
            all_bars[sym]["Close"].get(today, all_bars[sym]["Close"].iloc[-1]) * pos["shares"]
            for sym, pos in open_positions.items()
            if sym in all_bars
        )
        equity = cash + pos_value
        equity_curve[today] = equity

    return trade_log, pd.Series(equity_curve)


# ──────────────────────────────────────────────────────────────────────────
# PERFORMANCE STATS
# ──────────────────────────────────────────────────────────────────────────
def compute_performance_stats(trade_log: list, equity_curve: pd.Series,
                               starting_capital: float, years: int) -> dict:
    """Compute comprehensive performance statistics from trade log + equity curve."""
    if not trade_log:
        return {"error": "No trades taken"}

    df = pd.DataFrame(trade_log)

    # Filter to closed trades only (excludes partial exits that re-opened)
    closed = df[df["exit_reason"] != "partial_exit"].copy() if len(df) > 0 else df

    total_trades = len(closed)
    wins = closed[closed["pnl_dollars"] > 0]
    losses = closed[closed["pnl_dollars"] <= 0]

    win_rate = len(wins) / total_trades * 100 if total_trades > 0 else 0
    avg_win = wins["pnl_pct"].mean() if len(wins) > 0 else 0
    avg_loss = losses["pnl_pct"].mean() if len(losses) > 0 else 0
    gross_profit = wins["pnl_dollars"].sum()
    gross_loss = abs(losses["pnl_dollars"].sum())
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float("inf")
    avg_hold = closed["days_held"].mean() if len(closed) > 0 else 0

    # Equity curve stats
    final_equity = equity_curve.iloc[-1]
    total_return = (final_equity / starting_capital - 1) * 100
    cagr = ((final_equity / starting_capital) ** (1 / years) - 1) * 100 if years > 0 else 0

    # Sharpe ratio (daily returns)
    daily_returns = equity_curve.pct_change().dropna()
    sharpe = (daily_returns.mean() / daily_returns.std() * np.sqrt(252)) if daily_returns.std() > 0 else 0

    # Max drawdown
    roll_max = equity_curve.cummax()
    drawdown = (equity_curve - roll_max) / roll_max * 100
    max_drawdown = drawdown.min()

    # By exit reason
    reason_stats = df.groupby("exit_reason")["pnl_dollars"].agg(["count", "sum", "mean"]).round(2)

    # Top 10 symbols by P&L
    top10 = df.groupby("symbol")["pnl_dollars"].sum().sort_values(ascending=False).head(10)

    return {
        "total_trades": total_trades,
        "win_rate_pct": round(win_rate, 1),
        "avg_win_pct": round(avg_win, 2),
        "avg_loss_pct": round(avg_loss, 2),
        "profit_factor": round(profit_factor, 2),
        "gross_profit": round(gross_profit, 2),
        "gross_loss": round(gross_loss, 2),
        "net_pnl": round(gross_profit - gross_loss, 2),
        "avg_hold_days": round(avg_hold, 1),
        "starting_capital": starting_capital,
        "final_equity": round(final_equity, 2),
        "total_return_pct": round(total_return, 2),
        "cagr_pct": round(cagr, 2),
        "sharpe_ratio": round(sharpe, 2),
        "max_drawdown_pct": round(max_drawdown, 2),
        "reason_stats": reason_stats,
        "top10_symbols": top10,
    }


def print_stats(stats: dict):
    """Pretty-print performance statistics."""
    if "error" in stats:
        print(f"\n[RESULTS] {stats['error']}")
        return

    print("\n" + "=" * 60)
    print("  QULLAMAGGIE BACKTEST RESULTS")
    print("=" * 60)

    summary = [
        ["Total Trades", stats["total_trades"]],
        ["Win Rate", f"{stats['win_rate_pct']}%"],
        ["Avg Win", f"+{stats['avg_win_pct']}%"],
        ["Avg Loss", f"{stats['avg_loss_pct']}%"],
        ["Profit Factor", stats["profit_factor"]],
        ["Gross Profit", f"${stats['gross_profit']:,.2f}"],
        ["Gross Loss", f"${stats['gross_loss']:,.2f}"],
        ["Net P&L", f"${stats['net_pnl']:,.2f}"],
        ["Avg Hold (days)", stats["avg_hold_days"]],
        ["", ""],
        ["Starting Capital", f"${stats['starting_capital']:,.2f}"],
        ["Final Equity", f"${stats['final_equity']:,.2f}"],
        ["Total Return", f"{stats['total_return_pct']}%"],
        ["CAGR", f"{stats['cagr_pct']}%"],
        ["Sharpe Ratio", stats["sharpe_ratio"]],
        ["Max Drawdown", f"{stats['max_drawdown_pct']}%"],
    ]

    print(tabulate(summary, tablefmt="rounded_outline"))

    print("\n  BY EXIT REASON:")
    print(stats["reason_stats"].to_string())

    print("\n  TOP 10 SYMBOLS BY P&L:")
    for sym, pnl in stats["top10_symbols"].items():
        print(f"    {sym:<8} ${pnl:>10,.2f}")

    print("=" * 60)


# ──────────────────────────────────────────────────────────────────────────
# EXPORT TO EXCEL
# ──────────────────────────────────────────────────────────────────────────
def export_to_excel(trade_log: list, equity_curve: pd.Series, stats: dict, output_path: str):
    """Export trade log, equity curve, and stats to a formatted Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("[WARN] openpyxl not installed — saving CSV instead")
        pd.DataFrame(trade_log).to_csv(output_path.replace(".xlsx", ".csv"), index=False)
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Trade Log ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Trade Log"

    df_trades = pd.DataFrame(trade_log)
    if not df_trades.empty:
        # Column order
        cols = [
            "symbol", "entry_date", "exit_date", "days_held",
            "entry_price", "exit_price", "shares",
            "pnl_dollars", "pnl_pct", "exit_reason",
            "setup_score", "adr_pct", "consol_days",
            "consol_range_pct", "volume_ratio", "breakout_level",
            "prior_move_pct", "initial_risk_pct",
        ]
        cols = [c for c in cols if c in df_trades.columns]
        df_trades = df_trades[cols]

        # Header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(df_trades.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(df_trades.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                # Color P&L rows
                if df_trades.columns[col_idx - 1] == "pnl_dollars":
                    if isinstance(val, (int, float)) and val > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif isinstance(val, (int, float)) and val < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Auto-width columns
        for col in ws1.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    # ── Sheet 2: Equity Curve ──────────────────────────────────────
    ws2 = wb.create_sheet("Equity Curve")
    ws2.cell(1, 1, "Date").font = Font(bold=True)
    ws2.cell(1, 2, "Equity").font = Font(bold=True)
    for i, (date, val) in enumerate(equity_curve.items(), 2):
        ws2.cell(i, 1, str(date.date()) if hasattr(date, "date") else str(date))
        ws2.cell(i, 2, round(val, 2))

    # ── Sheet 3: Performance Summary ──────────────────────────────
    ws3 = wb.create_sheet("Performance Summary")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 20

    title_font = Font(bold=True, size=14)
    ws3.cell(1, 1, "QULLAMAGGIE BACKTEST — PERFORMANCE SUMMARY").font = title_font

    summary_rows = [
        ("", ""),
        ("TRADE STATISTICS", ""),
        ("Total Trades", stats.get("total_trades", 0)),
        ("Win Rate", f"{stats.get('win_rate_pct', 0)}%"),
        ("Avg Win %", f"+{stats.get('avg_win_pct', 0)}%"),
        ("Avg Loss %", f"{stats.get('avg_loss_pct', 0)}%"),
        ("Profit Factor", stats.get("profit_factor", 0)),
        ("Gross Profit", f"${stats.get('gross_profit', 0):,.2f}"),
        ("Gross Loss", f"${stats.get('gross_loss', 0):,.2f}"),
        ("Net P&L", f"${stats.get('net_pnl', 0):,.2f}"),
        ("Avg Hold (days)", stats.get("avg_hold_days", 0)),
        ("", ""),
        ("PORTFOLIO STATISTICS", ""),
        ("Starting Capital", f"${stats.get('starting_capital', 0):,.2f}"),
        ("Final Equity", f"${stats.get('final_equity', 0):,.2f}"),
        ("Total Return", f"{stats.get('total_return_pct', 0)}%"),
        ("CAGR", f"{stats.get('cagr_pct', 0)}%"),
        ("Sharpe Ratio", stats.get("sharpe_ratio", 0)),
        ("Max Drawdown", f"{stats.get('max_drawdown_pct', 0)}%"),
    ]

    for row_i, (label, value) in enumerate(summary_rows, 3):
        cell_a = ws3.cell(row_i, 1, label)
        cell_b = ws3.cell(row_i, 2, value)
        if label in ("TRADE STATISTICS", "PORTFOLIO STATISTICS"):
            cell_a.font = Font(bold=True)
            cell_a.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell_b.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    wb.save(output_path)
    print(f"\n[OUTPUT] Excel saved → {output_path}")


# ──────────────────────────────────────────────────────────────────────────
# ENTRYPOINT
# ──────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Qullamaggie Walk-Forward Backtest")
    parser.add_argument("--years", type=int, default=3, help="Years of history to backtest (default: 3)")
    parser.add_argument("--capital", type=float, default=100_000, help="Starting capital (default: 100000)")
    parser.add_argument("--symbol", type=str, default=None, help="Single symbol to backtest")
    parser.add_argument("--output", type=str, default="/root/.openclaw/workspace/.tmp/backtest_trades.xlsx",
                        help="Output Excel path")
    parser.add_argument("--refresh", action="store_true", help="Force refresh Nasdaq universe cache")
    parser.add_argument("--limit", type=int, default=0, help="Limit universe to N symbols (0 = all)")
    args = parser.parse_args()

    BT_CONFIG["starting_capital"] = args.capital

    print("\n" + "=" * 60)
    print("  QULLAMAGGIE WALK-FORWARD BACKTESTER")
    print(f"  {datetime.now(ET).strftime('%Y-%m-%d %H:%M ET')}")
    print("=" * 60)

    # Build universe
    if args.symbol:
        symbols = [args.symbol.upper(), "SPY"]
    else:
        symbols = fetch_nasdaq_universe(force_refresh=args.refresh)
        if not symbols:
            # Ultimate fallback: use scanner's built-in list
            print("[WARN] Universe empty — using built-in watchlist")
            from qullamaggie_scanner import CONFIG as SC
            symbols = []  # scanner doesn't export it, use a short list
        symbols = list(set(symbols + ["SPY"]))  # always include SPY for market filter
        if args.limit > 0:
            # Keep SPY and take `limit` from the rest
            others = [s for s in symbols if s != "SPY"][:args.limit]
            symbols = others + ["SPY"]

    print(f"[UNIVERSE] {len(symbols)} symbols to scan")

    # Download bars
    all_bars = download_all_bars(symbols, years=args.years)

    if not all_bars:
        print("[ERROR] No data downloaded. Check internet connection.")
        return

    # Run backtest
    trade_log, equity_curve = run_backtest(all_bars, args.capital, args.years)

    # Stats
    stats = compute_performance_stats(trade_log, equity_curve, args.capital, args.years)
    print_stats(stats)

    # Export
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    export_to_excel(trade_log, equity_curve, stats, args.output)

    print(f"\n✅ Done. Trade log: {args.output}")
    print(f"   Total trades in log: {len(trade_log)}")


if __name__ == "__main__":
    main()
