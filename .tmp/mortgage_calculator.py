import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mortgage Calculator"

# --- Colors ---
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "D6E4F0"
GREEN       = "E2EFDA"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"

def hfill(color):
    return PatternFill("solid", fgColor=color)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(white=True, size=11):
    return Font(name="Calibri", bold=True, color="FFFFFF" if white else "1F3864", size=size)

def body_font(bold=False):
    return Font(name="Calibri", bold=bold, size=10)

center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center")
right  = Alignment(horizontal="right",  vertical="center")

# ─────────────────────────────────────────────
# SECTION 1 — INPUT PANEL (rows 1-10)
# ─────────────────────────────────────────────

# Title bar
ws.merge_cells("A1:G1")
ws["A1"] = "🏠  Mortgage Calculator"
ws["A1"].font      = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
ws["A1"].fill      = hfill(BLUE_DARK)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# Sub-header
ws.merge_cells("A2:G2")
ws["A2"] = "Enter your loan details below — the table and summary update automatically"
ws["A2"].font      = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
ws["A2"].fill      = hfill(BLUE_MID)
ws["A2"].alignment = center
ws.row_dimensions[2].height = 18

# Input labels + named cells
inputs = [
    (4, "Home Price ($)",        500000,   "B4",  "#,##0.00"),
    (5, "Down Payment ($)",      100000,   "B5",  "#,##0.00"),
    (6, "Annual Interest Rate (%)", 6.5,   "B6",  "0.00%"),
    (7, "Loan Term (Years)",     30,       "B7",  "0"),
    (8, "Property Tax / Year ($)", 4800,   "B8",  "#,##0.00"),
    (9, "Insurance / Year ($)",  1200,     "B9",  "#,##0.00"),
]

ws["A3"] = "LOAN INPUTS"
ws["A3"].font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
ws["A3"].fill      = hfill(BLUE_MID)
ws["A3"].alignment = center
ws.merge_cells("A3:B3")

ws["D3"] = "QUICK SUMMARY"
ws["D3"].font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
ws["D3"].fill      = hfill(BLUE_MID)
ws["D3"].alignment = center
ws.merge_cells("D3:G3")

for row, label, val, cell, fmt in inputs:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font      = body_font(bold=True)
    ws[f"A{row}"].fill      = hfill(BLUE_LIGHT)
    ws[f"A{row}"].alignment = left
    ws[f"A{row}"].border    = border()

    ws[cell] = val
    if fmt == "0.00%":
        ws[cell].value = val / 100   # store as decimal for % format
    ws[cell].number_format = fmt
    ws[cell].fill      = hfill(WHITE)
    ws[cell].alignment = right
    ws[cell].border    = border()
    ws[cell].font      = body_font()

# Named ranges for formula references (use direct cell refs)
# B4 = Home Price, B5 = Down Payment, B6 = Rate (decimal), B7 = Term yrs, B8 = Tax/yr, B9 = Ins/yr

# Derived: Loan Amount, Monthly P&I, Total Payment, Total Interest
summary = [
    (4,  "Loan Amount",           "=B4-B5",                              "#,##0.00"),
    (5,  "Monthly P&I Payment",   '=IFERROR(PMT(B6/12,B7*12,-(B4-B5)),0)', "#,##0.00"),
    (6,  "Monthly Tax + Insurance",'=(B8+B9)/12',                        "#,##0.00"),
    (7,  "Total Monthly Payment",  '=D5+D6',                             "#,##0.00"),
    (8,  "Total Paid (Principal+Interest)", '=D5*B7*12',                "#,##0.00"),
    (9,  "Total Interest Paid",    '=D8-(B4-B5)',                        "#,##0.00"),
]

for row, label, formula, fmt in summary:
    ws[f"D{row}"] = label
    ws[f"D{row}"].font      = body_font(bold=True)
    ws[f"D{row}"].fill      = hfill(BLUE_LIGHT)
    ws[f"D{row}"].alignment = left
    ws[f"D{row}"].border    = border()
    ws.merge_cells(f"D{row}:F{row}")

    ws[f"G{row}"] = formula
    ws[f"G{row}"].number_format = fmt
    ws[f"G{row}"].fill      = hfill(GREEN)
    ws[f"G{row}"].alignment = right
    ws[f"G{row}"].border    = border()
    ws[f"G{row}"].font      = body_font(bold=(row in [7,9]))

# Blank spacer row
ws.row_dimensions[10].height = 8

# ─────────────────────────────────────────────
# SECTION 2 — AMORTIZATION TABLE (rows 12+)
# ─────────────────────────────────────────────

ws.merge_cells("A11:G11")
ws["A11"] = "📅  Full Amortization Schedule"
ws["A11"].font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
ws["A11"].fill      = hfill(BLUE_DARK)
ws["A11"].alignment = center
ws.row_dimensions[11].height = 24

# Table headers
headers = ["Payment #", "Month / Year", "Beginning Balance",
           "Principal", "Interest", "Ending Balance", "Cumulative Interest"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=12, column=col, value=h)
    c.font      = header_font()
    c.fill      = hfill(BLUE_MID)
    c.alignment = center
    c.border    = border()
ws.row_dimensions[12].height = 20

# Amortization rows — use formulas so they respond to input changes
# Row 13 = payment 1
# We'll generate 360 rows (30-year max) with IF guards so unused rows are blank

START = 13
MAX_PAYMENTS = 360   # 30 years × 12

import datetime
base_date = datetime.date(2026, 4, 1)  # first payment month

for i in range(MAX_PAYMENTS):
    r = START + i
    n = i + 1   # payment number

    # Toggle fill for zebra striping
    fill = hfill(WHITE) if i % 2 == 0 else hfill(GREY)

    # Payment # — only show if n <= B7*12
    ws.cell(r, 1).value          = f"=IF({n}<=B$7*12,{n},\"\")"
    ws.cell(r, 1).alignment      = center
    ws.cell(r, 1).fill           = fill
    ws.cell(r, 1).border         = border()
    ws.cell(r, 1).font           = body_font()

    # Month / Year
    pay_date = base_date.replace(month=((base_date.month - 1 + i) % 12) + 1,
                                  year=base_date.year + (base_date.month - 1 + i) // 12)
    ws.cell(r, 2).value          = f'=IF({n}<=B$7*12,TEXT(DATE({pay_date.year},{pay_date.month},1),"mmm yyyy"),"")'
    ws.cell(r, 2).alignment      = center
    ws.cell(r, 2).fill           = fill
    ws.cell(r, 2).border         = border()
    ws.cell(r, 2).font           = body_font()

    # Beginning Balance
    if i == 0:
        beg_bal = "=IF(1<=B$7*12,B4-B5,\"\")"
    else:
        beg_bal = f'=IF({n}<=B$7*12,F{r-1},"")'
    ws.cell(r, 3).value          = beg_bal
    ws.cell(r, 3).number_format  = "#,##0.00"
    ws.cell(r, 3).alignment      = right
    ws.cell(r, 3).fill           = fill
    ws.cell(r, 3).border         = border()
    ws.cell(r, 3).font           = body_font()

    # Interest portion
    ws.cell(r, 5).value          = f'=IF({n}<=B$7*12,C{r}*B$6/12,"")'
    ws.cell(r, 5).number_format  = "#,##0.00"
    ws.cell(r, 5).alignment      = right
    ws.cell(r, 5).fill           = fill
    ws.cell(r, 5).border         = border()
    ws.cell(r, 5).font           = body_font()

    # Principal portion
    ws.cell(r, 4).value          = f'=IF({n}<=B$7*12,IFERROR(PMT(B$6/12,B$7*12,-(B4-B5)),0)-E{r},"")'
    ws.cell(r, 4).number_format  = "#,##0.00"
    ws.cell(r, 4).alignment      = right
    ws.cell(r, 4).fill           = fill
    ws.cell(r, 4).border         = border()
    ws.cell(r, 4).font           = body_font()

    # Ending Balance
    ws.cell(r, 6).value          = f'=IF({n}<=B$7*12,MAX(C{r}-D{r},0),"")'
    ws.cell(r, 6).number_format  = "#,##0.00"
    ws.cell(r, 6).alignment      = right
    ws.cell(r, 6).fill           = fill
    ws.cell(r, 6).border         = border()
    ws.cell(r, 6).font           = body_font()

    # Cumulative Interest
    if i == 0:
        ws.cell(r, 7).value = f'=IF({n}<=B$7*12,E{r},"")'
    else:
        ws.cell(r, 7).value = f'=IF({n}<=B$7*12,G{r-1}+E{r},"")'
    ws.cell(r, 7).number_format  = "#,##0.00"
    ws.cell(r, 7).alignment      = right
    ws.cell(r, 7).fill           = fill
    ws.cell(r, 7).border         = border()
    ws.cell(r, 7).font           = body_font()

# ─────────────────────────────────────────────
# Column widths
# ─────────────────────────────────────────────
col_widths = [12, 14, 20, 16, 16, 18, 20]
for col, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Freeze panes below header row of amortization table
ws.freeze_panes = "A13"

# ─────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────
out = "/root/.openclaw/workspace/.tmp/Mortgage_Calculator.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved → {out}")
